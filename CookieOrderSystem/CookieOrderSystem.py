from re import M
from flask import Flask, render_template, request, jsonify
#flask: 建立應用模組
#render_template : 渲染html模板
#request: 取得前端發送資料
#jsonify: 將資料轉為JSON格式回傳
 
from openpyxl import load_workbook, Workbook
#匯入openpyxl 模組以操作 exel(.xlsx)

import os
#操作檔案與路徑

import datetime

app = Flask(__name__)
#建立應用模組 __name__代表目前模組名稱

#設定Excel檔案路徑
DATA_FOLDER=os.path.join(os.getcwd(), 'data')
COOKIE_FILE = os.path.join(DATA_FOLDER, "cookie.xlsx")
ORDERS_FILE = os.path.join(DATA_FOLDER, "orders.xlsx")

def read_cookies():
    wb = load_workbook(COOKIE_FILE)
    ws = wb.active
    cookies =[]
    #開啟 cookies 檔並使用預設工作表。

    cookies = []

    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value for cell in row]
        cookie = {
            "ID": row[0].value,
            "名稱": row[1].value,
            "圖片檔名": row[2].value.strip(),
            "價格": row[3].value
            #用 .value 才能取出裡面真正的字串、數字
        }
        cookies.append(cookie)
        #每列資料組成字典，再加入cookies清單中

    return cookies


def append_order(order_items):
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["訂單編號", "日期", "餅乾ID", "數量", "總金額"])
        #若訂單檔不存在，則建立新檔

    else:
        wb = load_workbook(ORDERS_FILE)
        ws = wb.active
        #if 檔案存在，則開啟檔案

    order_id = int(datetime.datetime.now().timestamp())
    order_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    #格式化時間做為訂單時間

    for item in order_items:
        ws.append([
            order_id, 
            order_date, 
            item["cookieID"],
            item["quantity"], 
            item['total']])
        #將訂單資料加入訂單檔

    wb.save(ORDERS_FILE)
    return order_id
    #save file


@app.route('/')
#when user visits the website
def index():
    cookies = read_cookies()
    #call read_cookies function
    return render_template("index.html", cookies=cookies)
    #render index.html and pass cookies data to it

@app.route('/confirm', methods=["POST"])
def confirm_order():
    data = request.get_json()
    #get data JSON from request

    order_items = data.get('order',[])
    order_id = append_order(order_items)
    #call append_order function write order data to file

    return jsonify({"order_id": order_id})
    #return order_id in JSON format

if __name__ == "__main__":
    # 開啟除錯模式以便開發測試
    app.run(debug=True)
    #啟動伺服器, open Flask